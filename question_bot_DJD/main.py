from data import db_session
from data.applications import Applic
from data.admins import Admin
from data.answers import Answer
import telebot
import random
from telebot import types
from openpyxl import Workbook
from openpyxl.styles import Font
from get_theme import get_theme

db_session.global_init("db/DJD_bot.db")

token = "SECRET TOKEN"
bot = telebot.TeleBot(token)

ALPHABET = "qwertyuiopasdfghjklzxcvbnm"
DIGITAL = "1234567890"


@bot.message_handler(commands=["start"])
def start_message(message):
    bot.send_message(
        message.chat.id,
        "Здравствуйте! Перед вами бот, который ответит на ваши интересующие вопросы.",
        reply_markup=types.ReplyKeyboardRemove()
    )
    c1 = types.BotCommand(command="start", description="Старт бота")
    c2 = types.BotCommand(command="admin", description="Вход в админ панель")
    bot.set_my_commands([c1, c2])
    bot.set_chat_menu_button(message.chat.id, types.MenuButtonCommands("commands"))


@bot.message_handler(commands=["admin"])
def admin_message(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Назад")
    markup.add(item1)
    bot.send_message(
        message.chat.id, "Пожалуйста предоставте ваш код доступа", reply_markup=markup
    )
    bot.register_next_step_handler(message, password)


def password(message):
    if message.text == "Назад":
        bot.send_message(
            message.chat.id,
            "Вы вернулись назад",
            reply_markup=types.ReplyKeyboardRemove(),
        )
        bot.register_next_step_handler(message, button_message)
    else:
        db_sess = db_session.create_session()
        admins = db_sess.query(Admin).all()
        admin = None
        for i in admins:
            if i and i.check_password(message.text):
                admin = i
                break
        if admin and admin.login == "MainPerson":
            bot.send_message(message.chat.id, "Здравствуйте! Главный администратор")
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item1 = types.KeyboardButton("Добавить нового администратора")
            item2 = types.KeyboardButton("Удалить администратора")
            item3 = types.KeyboardButton("Список администраторов")
            item4 = types.KeyboardButton("Вопросы пользователей")
            item5 = types.KeyboardButton("Назад")
            items = [item1, item2, item3, item4, item5]
            for item in items:
                markup.add(item)
            bot.send_message(message.chat.id, "Выберите действие", reply_markup=markup)
            bot.register_next_step_handler(message, functions_main_admin)
        elif admin:
            bot.send_message(message.chat.id, f"Здравствуйте, {admin.login}!")
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item1 = types.KeyboardButton("Вопросы пользователей")
            item2 = types.KeyboardButton("Назад")
            markup.add(item1)
            markup.add(item2)
            bot.send_message(message.chat.id, "Выберите действие", reply_markup=markup)
            bot.register_next_step_handler(message, function_admin)
        else:
            bot.send_message(message.chat.id, "Веденный код неверный")
            bot.register_next_step_handler(message, password)


def functions_main_admin(message):
    if message.text == "Добавить нового администратора":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Отмена")
        markup.add(item1)
        msg = bot.send_message(
            message.chat.id, "Введите имя нового администратора", reply_markup=markup
        )
        bot.register_next_step_handler(msg, new_admin)
    elif message.text == "Удалить администратора":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Отмена")
        markup.add(item1)
        msg = bot.send_message(
            message.chat.id, "Введите имя администратора", reply_markup=markup
        )
        bot.register_next_step_handler(msg, delete_admin)
    elif message.text == "Список администраторов":
        bot.send_message(message.chat.id, "Сбор данных...")
        db_sess = db_session.create_session()
        admins = db_sess.query(Admin).all()
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "id"
        ws["B1"] = "Логин"
        for ind, i in enumerate(admins):
            ws[f"A{ind + 2}"] = i.id
            ws[f"B{ind + 2}"] = i.login
        wb.save("question_bot_DJD/files/admins.xlsx")
        with open("question_bot_DJD/files/admins.xlsx", "rb") as f:
            bot.send_document(message.chat.id, f)
        bot.send_message(message.chat.id, "Выберите действие")
        bot.register_next_step_handler(message, functions_main_admin)
    elif message.text == "Вопросы пользователей":
        bot.send_message(message.chat.id, "Сбор данных...")
        db_sess = db_session.create_session()
        applics = db_sess.query(Applic).all()
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "id"
        ws["B1"] = "Тема"
        ws["C1"] = "Текст"
        ws["D1"] = "Дата"
        ws["E1"] = "Статус"
        ws["F1"] = "id чата"
        for ind, i in enumerate(applics):
            ws[f"A{ind + 2}"] = i.id
            ws[f"F{ind + 2}"] = i.message_id
            ws[f"B{ind + 2}"] = i.theme
            ws[f"C{ind + 2}"] = i.text
            ws[f"D{ind + 2}"] = str(i.created_date)
            if i.status:
                ws[f"E{ind + 2}"] = "Готово"
                ws[f"E{ind + 2}"].font = Font(color="0000FF00")
            else:
                ws[f"E{ind + 2}"] = "Не готово"
                ws[f"E{ind + 2}"].font = Font(color="00FF0000")
        wb.save("question_bot_DJD/files/applications.xlsx")
        with open("question_bot_DJD/files/applications.xlsx", "rb") as f:
            bot.send_document(message.chat.id, f)
        bot.send_message(message.chat.id, "Выберите действие")
        bot.register_next_step_handler(message, functions_main_admin)
    elif message.text == "Назад":
        bot.send_message(
            message.chat.id,
            "Введите интересующий вопрос",
            reply_markup=types.ReplyKeyboardRemove(),
        )
        bot.register_next_step_handler(message, button_message)
    else:
        bot.send_message(message.chat.id, "Такой команды не существует")
        bot.register_next_step_handler(message, functions_main_admin)


def new_admin(message):
    if message.text == "Отмена":
        bot.send_message(message.chat.id, "Вы вернулись назад")
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Добавить нового администратора")
        item2 = types.KeyboardButton("Удалить администратора")
        item3 = types.KeyboardButton("Список администраторов")
        item4 = types.KeyboardButton("Вопросы пользователей")
        item5 = types.KeyboardButton("Назад")
        items = [item1, item2, item3, item4, item5]
        for item in items:
            markup.add(item)
        bot.send_message(message.chat.id, "Выберите действие", reply_markup=markup)
        bot.register_next_step_handler(message, functions_main_admin)
    else:
        bot.send_message(message.chat.id, "Добовляется новый администратор...")
        passw = random.choices(ALPHABET, k=7)
        passw += random.choices(DIGITAL, k=3)
        passw = "".join(passw)
        db_sess = db_session.create_session()
        admin = Admin(login=message.text)
        admin.set_password(passw)
        db_sess.add(admin)
        db_sess.commit()
        bot.send_message(
            message.chat.id,
            f"Добавлен новый администратор:\n Логин: {message.text}\n Пароль: {passw}\nВнимание! Пароли не хранятся в чистом виде в целях безопасности. Вы не сможете больше получить доступ к этому паролю.",
        )
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Добавить нового администратора")
        item2 = types.KeyboardButton("Удалить администратора")
        item3 = types.KeyboardButton("Список администраторов")
        item4 = types.KeyboardButton("Вопросы пользователей")
        item5 = types.KeyboardButton("Назад")
        items = [item1, item2, item3, item4, item5]
        for item in items:
            markup.add(item)
        bot.send_message(message.chat.id, "Выберите действие", reply_markup=markup)
        bot.register_next_step_handler(message, functions_main_admin)


def delete_admin(message):
    if message.text == "Отмена":
        bot.send_message(message.chat.id, "Вы вернулись назад")
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Добавить нового администратора")
        item2 = types.KeyboardButton("Удалить администратора")
        item3 = types.KeyboardButton("Список администраторов")
        item4 = types.KeyboardButton("Вопросы пользователей")
        item5 = types.KeyboardButton("Назад")
        items = [item1, item2, item3, item4, item5]
        for item in items:
            markup.add(item)
        bot.send_message(message.chat.id, "Выберите действие", reply_markup=markup)
        bot.register_next_step_handler(message, functions_main_admin)
    else:
        bot.send_message(message.chat.id, "Происходит удаление админа...")
        if message.text == "MainPerson":
            bot.send_message(message.chat.id, "Нельзя удалить главного администратора")
            bot.register_next_step_handler(message, delete_admin)
            return
        db_sess = db_session.create_session()
        del_admin = db_sess.query(Admin).filter(Admin.login == message.text).first()
        if del_admin:
            db_sess.delete(del_admin)
            db_sess.commit()

            bot.send_message(
                message.chat.id, "Администратор успешно удален из базы данных"
            )
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item1 = types.KeyboardButton("Добавить нового администратора")
            item2 = types.KeyboardButton("Удалить администратора")
            item3 = types.KeyboardButton("Список администраторов")
            item4 = types.KeyboardButton("Вопросы пользователей")
            item5 = types.KeyboardButton("Назад")
            items = [item1, item2, item3, item4, item5]
            for item in items:
                markup.add(item)
            bot.send_message(message.chat.id, "Выберите действие", reply_markup=markup)
            bot.register_next_step_handler(message, functions_main_admin)
        else:
            bot.send_message(message.chat.id, "Администратор с таким логином не найден")
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item1 = types.KeyboardButton("Добавить нового администратора")
            item2 = types.KeyboardButton("Удалить администратора")
            item3 = types.KeyboardButton("Список администраторов")
            item4 = types.KeyboardButton("Вопросы пользователей")
            item5 = types.KeyboardButton("Назад")
            items = [item1, item2, item3, item4, item5]
            for item in items:
                markup.add(item)
            bot.send_message(message.chat.id, "Выберите действие", reply_markup=markup)
            bot.register_next_step_handler(message, functions_main_admin)


def function_admin(message):
    if message.text == "Назад":
        bot.send_message(
            message.chat.id,
            "Введите интересующий вопрос",
            reply_markup=types.ReplyKeyboardRemove(),
        )
        bot.register_next_step_handler(message, button_message)
    elif message.text == "Вопросы пользователей":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        bot.send_message(
            message.chat.id, "Сбор данных...", reply_markup=types.ReplyKeyboardRemove()
        )
        db_sess = db_session.create_session()
        applics = db_sess.query(Applic).filter(Applic.status == 0).all()
        text = "Вопросы:\n"
        for app in applics:
            text += f"{app.id}) Тема: {app.theme}\n  Вопрос: {app.text}\n"
            markup.add(types.KeyboardButton(str(app.id)))
        markup.add(types.KeyboardButton("Отмена"))
        bot.send_message(message.chat.id, text, reply_markup=markup)
        bot.register_next_step_handler(message, questions)
    else:
        bot.send_message(message.chat.id, "Такой команды не существует")
        bot.register_next_step_handler(message, function_admin)


def questions(message):
    global applic_id

    if message.text == "Отмена":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Вопросы пользователей")
        item2 = types.KeyboardButton("Назад")
        markup.add(item1)
        markup.add(item2)
        bot.send_message(message.chat.id, "Выберите действие", reply_markup=markup)
        bot.register_next_step_handler(message, function_admin)
    else:
        applic_id = int(message.text)
        db_sess = db_session.create_session()
        question = db_sess.query(Applic).filter(Applic.id == int(message.text)).first()
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton("Отмена"))
        if question:
            bot.send_message(
                message.chat.id,
                f"Ответьте на этот вопрос: {question.text}",
                reply_markup=markup,
            )
            bot.register_next_step_handler(message, answer)
        else:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item1 = types.KeyboardButton("Вопросы пользователей")
            item2 = types.KeyboardButton("Назад")
            markup.add(item1)
            markup.add(item2)
            bot.send_message(
                message.chat.id,
                "По неизвестной причине такой вопрос не найден",
                reply_markup=markup,
            )
            bot.register_next_step_handler(message, function_admin)


def answer(message):
    global applic_id

    if message.text == "Отмена":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        bot.send_message(
            message.chat.id, "Сбор данных", reply_markup=types.ReplyKeyboardRemove()
        )
        db_sess = db_session.create_session()
        applics = db_sess.query(Applic).filter(Applic.status == 1).all()
        text = "Вопросы:\n"
        for app in applics:
            text += f"{app.id}) Тема: {app.theme}\n  Вопрос: {app.text}\n"
            markup.add(types.KeyboardButton(str(app.id)))
        markup.add(types.KeyboardButton("Отмена"))
        bot.send_message(message.chat.id, text, reply_markup=markup)
        bot.register_next_step_handler(message, questions)
    else:
        bot.send_message(message.chat.id, "Ответ обрабатывается...")
        db_sess = db_session.create_session()
        ans = Answer(text=message.text, applic_id=applic_id)
        db_sess.add(ans)
        bot.send_message(message.chat.id, "Спасибо за ответ!")
        answer_user(ans, message)
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        bot.send_message(
            message.chat.id, "Сбор данных", reply_markup=types.ReplyKeyboardRemove()
        )
        db_sess = db_session.create_session()
        applics = db_sess.query(Applic).filter(Applic.status == 0).all()
        text = "Вопросы:\n"
        for app in applics:
            text += f"{app.id}) Тема: {app.theme}\n  Вопрос: {app.text}\n"
            markup.add(types.KeyboardButton(str(app.id)))
        markup.add(types.KeyboardButton("Отмена"))
        bot.send_message(message.chat.id, text, reply_markup=markup)
        bot.register_next_step_handler(message, questions)


def answer_user(ans: "Answer", message):
    db_sess = db_session.create_session()
    applic = db_sess.query(Applic).filter(Applic.id == ans.applic_id).first()
    bot.send_message(applic.message_id, "Вам пришел ответ от преподователя!")
    bot.send_message(applic.message_id, ans.text)
    applic.status = True
    db_sess.merge(applic)
    db_sess.commit()


@bot.message_handler(content_types="text")
def button_message(message):
    global text
    global them
    text = message.text
    maximum = 0
    # items = [i[1] for i in get_theme(message.text).items()]
    them = []
    all = 0
    for item in get_theme(message.text).items():
        all += item[1]
        if item[1] > maximum:
            maximum = item[1]
    for item in get_theme(message.text).items():
        if item[1] == maximum and maximum != 0:
            them.append(item[0])
    if all > 0:
        bot.send_message(
            message.chat.id,
            f"""Темы:
1. Прокат - {get_theme(message.text)["прокат"] / all * 100}%
2. Экскурсия -  {get_theme(message.text)['экскурсия'] / all * 100}%
3. Запись - {get_theme(message.text)["запись"] / all * 100}%
4. Расписание - {get_theme(message.text)["расписание"] / all * 100}%
5. Возраст -  {get_theme(message.text)["возраст"] / all * 100}%
6. Местоположение - {get_theme(message.text)["местоположение"] / all * 100}%""",
        )
    flag = False
    flag2 = False
    if "прокат" in them:
        bot.send_media_group(
            message.chat.id,
            [
                types.InputMediaPhoto(
                    open("question_bot_DJD/files/1smena.jpg", "rb"), caption=""
                ),
                types.InputMediaPhoto(
                    open("question_bot_DJD/files/2smena.jpg", "rb"), caption=""
                ),
                types.InputMediaPhoto(
                    open("question_bot_DJD/files/tarif.jpg", "rb"), caption=""
                ),
                types.InputMediaPhoto(
                    open("question_bot_DJD/files/work.jpg", "rb"), caption=""
                ),
            ],
        )
        bot.send_message(message.chat.id, "Семейная поездка или групповая?")
        flag = True
        bot.register_next_step_handler(message, rent)
        flag2 = True
    if "экскурсия" in them:
        flag = True
        bot.send_message(message.chat.id, "ответ на тему экскурсия")
    if "запись" in them:
        flag = True
        bot.send_message(
            message.chat.id,
            """✏ Для записи на обучение Вам необходимо:

💼 Собрать пакет документов
📎 копия документа удостоверяющего личность ребенка
📎 копия СНИЛС ребенка
📎 4 цветные фотографии 3х4 без уголка
📎 заполнить заявление 👇🏼*
📎 заполнить согласие на обратку персональных данных 👇🏼

📬 принести пакет документов по адресу Чита, 1-я Ипподромная 5 Читинская детская железная дорога

📆 выбрать удобный день для посещения занятий

💬 ожидать сообщения о начале занятий

🚀 быть готовым вовремя прийти на занятие""",
        )
        bot.send_document(
            message.chat.id,
            open(
                "question_bot_DJD/files/Заявление_на_поступление_на_ЧДЖД+анкет+расписка+согласие_на_обр.pdf",
                "rb",
            ),
        )
    if "расписание" in them:
        flag = True
        bot.send_media_group(
            message.chat.id,
            [
                types.InputMediaPhoto(
                    open("question_bot_DJD/files/1smena.jpg", "rb"), caption=""
                ),
                types.InputMediaPhoto(
                    open("question_bot_DJD/files/2smena.jpg", "rb"), caption=""
                ),
                types.InputMediaPhoto(
                    open("question_bot_DJD/files/work.jpg", "rb"), caption=""
                ),
            ],
        )
    if "возраст" in them:
        flag = True
        bot.send_message(
            message.chat.id,
            "Мы берем на обучение детей с 5 класса. (9 и 11 класс не берем)",
        )
    if "местоположение" in them:
        flag = True
        bot.send_message(
            message.chat.id, "1-я ипподромная, 5. Остановки: Пожарка, Локомотив"
        )
    if not flag:
        bot.send_message(message.chat.id, "Я, пока что, не могу ответить на этот вопрос. Мне передать этот вопрос преподавателю?")
        bot.register_next_step_handler(message, final)
        flag2 = True
    elif not flag2:
        bot.send_message(message.chat.id, "Наш ответ вас устроил?")
        bot.register_next_step_handler(message, status)


def rent(message):
    if 'групповая' in message.text.lower():
        bot.send_message(message.chat.id, "групповая поездка")
        bot.send_message(message.chat.id, "Наш ответ вас устроил?")
        bot.register_next_step_handler(message, status)
    elif 'Семейная' in message.text.lower():
        bot.send_message(message.chat.id, "личная поездка")
        bot.send_message(message.chat.id, "Наш ответ вас устроил?")
        bot.register_next_step_handler(message, status)
    else:
        bot.send_message(message.chat.id, "Введите либо 'Cемейная' или 'Групповая'")
        bot.register_next_step_handler(message, rent)
    


def status(message):
    global text
    global them
    if "да" in message.text.lower():
        bot.send_message(message.chat.id, "Спасибо за отзыв")
        db_sess = db_session.create_session()
        applic = Applic(
            message_id=message.chat.id, text=text, theme=them[0], status=True
        )
        db_sess.add(applic)
        db_sess.commit()
    elif "нет" in message.text.lower():
        bot.send_message(message.chat.id, "Скоро преподаватели ответят на ваш вопрос")
        db_sess = db_session.create_session()
        db_sess.add(
            Applic(
                message_id=message.chat.id, text=text, theme=them[0], status=False
            )
        )
        db_sess.commit()
    else:
        bot.send_message(message.chat.id, "Ответьте просто 'да' или 'нет'")
        bot.register_next_step_handler(message, status)


def final(message):
    global text
    global them
    if "нет" in message.text.lower():
        bot.send_message(message.chat.id, "Выбор сделан")
    elif "да" in message.text.lower():
        bot.send_message(message.chat.id, "Скоро преподаватели ответят на ваш вопрос")
        db_sess = db_session.create_session()
        db_sess.add(
            Applic(
                message_id=message.chat.id, text=text, theme="Неизвестно", status=False
            )
        )
        db_sess.commit()
    else:
        bot.send_message(message.chat.id, "Ответьте просто 'да' или 'нет'")
        bot.register_next_step_handler(message, final)


bot.infinity_polling()
